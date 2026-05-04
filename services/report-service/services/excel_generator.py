"""
Excel Report Generator
Uses openpyxl to generate Excel reports
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import httpx
import logging
from datetime import datetime
from typing import Optional, Dict, Any, List

import sys
sys.path.append('/app')
from config.settings import get_settings

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generate Excel reports for players, teams, and matches"""

    def __init__(self):
        self.settings = get_settings()
        self.header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=12)
        self.title_font = Font(size=16, bold=True, color="1a237e")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    async def _get_json(self, url: str) -> Dict[str, Any]:
        try:
            async with httpx.AsyncClient(timeout=10.0) as client:
                response = await client.get(url)
                if response.status_code == 200:
                    payload = response.json()
                    if isinstance(payload, dict) and 'data' in payload:
                        return payload['data'] or {}
                    return payload if isinstance(payload, dict) else {}
        except Exception as e:
            logger.error(f"Error fetching {url}: {e}")
        return {}

    async def fetch_player_data(self, player_id: str) -> Dict[str, Any]:
        data = await self._get_json(f"{self.settings.player_service_url}/api/v2/players/{player_id}")
        return data or {"player_id": player_id, "name": "Unknown Player"}

    async def fetch_player_stats(self, player_id: str) -> Dict[str, Any]:
        data = await self._get_json(f"{self.settings.statistics_service_url}/api/v2/statistics/player/{player_id}")
        if isinstance(data, dict) and 'stats' in data:
            return data['stats']
        return data or {}

    async def fetch_team_data(self, team_id: str) -> Dict[str, Any]:
        data = await self._get_json(f"{self.settings.team_service_url}/api/v2/teams/{team_id}")
        return data or {"team_id": team_id, "name": "Unknown Team"}

    async def fetch_team_stats(self, team_id: str) -> Dict[str, Any]:
        data = await self._get_json(f"{self.settings.statistics_service_url}/api/v2/statistics/team/{team_id}")
        if isinstance(data, dict) and 'stats' in data:
            return data['stats']
        return data or {}

    async def fetch_match_data(self, match_id: str) -> Dict[str, Any]:
        data = await self._get_json(f"{self.settings.match_service_url}/api/v2/matches/{match_id}")
        return data or {"match_id": match_id}

    def _write_header_row(self, ws, row: int, columns: List[str]):
        for col, label in enumerate(columns, 1):
            cell = ws.cell(row=row, column=col, value=label)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border

    def _write_data_row(self, ws, row: int, values: List[Any]):
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = self.border

    def _save_workbook(self, wb: Workbook) -> bytes:
        buffer = io.BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()
        buffer.close()
        return content

    async def generate_player_report(self, player_id: str, include_stats: bool = True) -> bytes:
        player = await self.fetch_player_data(player_id)
        stats = await self.fetch_player_stats(player_id) if include_stats else {}

        wb = Workbook()
        ws_info = wb.active
        ws_info.title = "Player Information"

        ws_info['A1'] = f"Player Report: {player.get('name', 'Unknown')}"
        ws_info['A1'].font = self.title_font
        ws_info.merge_cells('A1:B1')

        ws_info['A3'] = "Generated:"
        ws_info['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        ws_info['A4'] = "Player ID:"
        ws_info['B4'] = player_id

        self._write_header_row(ws_info, 6, ['Field', 'Value'])
        info_fields = [
            ('Name', player.get('name') or 'N/A'),
            ('Position', player.get('position') or 'N/A'),
            ('Detailed Position', player.get('detailed_position') or player.get('detailedPosition') or 'N/A'),
            ('Team', player.get('club') or player.get('team_name') or player.get('teamName') or 'N/A'),
            ('Nationality', player.get('nationality') or 'N/A'),
            ('Age', player.get('age') or 'N/A'),
            ('Height', player.get('height') or 'N/A'),
            ('Weight', player.get('weight') or 'N/A'),
            ('Preferred Foot', player.get('preferred_foot') or player.get('preferredFoot') or 'N/A'),
        ]
        for row_idx, (field, value) in enumerate(info_fields, 7):
            self._write_data_row(ws_info, row_idx, [field, str(value)])

        ws_info.column_dimensions['A'].width = 25
        ws_info.column_dimensions['B'].width = 30

        if include_stats:
            ws_stats = wb.create_sheet("Statistics")
            ws_stats['A1'] = "Season Statistics"
            ws_stats['A1'].font = self.title_font
            ws_stats.merge_cells('A1:B1')

            self._write_header_row(ws_stats, 3, ['Statistic', 'Value'])
            stat_labels = {
                'goals': 'Goals',
                'goal_assist': 'Assists',
                'assists': 'Assists',
                'appearances': 'Appearances',
                'games_played': 'Appearances',
                'minutes_played': 'Minutes Played',
                'total_pass': 'Total Passes',
                'accurate_pass': 'Accurate Passes',
                'pass_accuracy': 'Pass Accuracy (%)',
                'total_scoring_att': 'Total Shots',
                'ontarget_scoring_att': 'Shots on Target',
                'yellow_card': 'Yellow Cards',
                'red_card': 'Red Cards',
                'total_tackle': 'Tackles',
                'interception': 'Interceptions',
            }
            seen_labels = set()
            row_idx = 4
            for key, label in stat_labels.items():
                if label in seen_labels:
                    continue
                val = stats.get(key)
                if val is not None:
                    self._write_data_row(ws_stats, row_idx, [label, val])
                    seen_labels.add(label)
                    row_idx += 1

            sb = player.get('statsbomb_enrichment') or player.get('statsbombEnrichment')
            if isinstance(sb, dict):
                row_idx += 1
                ws_stats.cell(row=row_idx, column=1, value="Advanced Metrics (StatsBomb)").font = Font(bold=True)
                row_idx += 1
                self._write_header_row(ws_stats, row_idx, ['Metric', 'Value'])
                row_idx += 1
                for key, label in [('total_xg', 'Total xG'), ('total_obv', 'Total OBV'),
                                    ('passes', 'Passes'), ('shots', 'Shots'), ('goals', 'Goals'),
                                    ('avg_pass_success_prob', 'Avg Pass Success %')]:
                    val = sb.get(key)
                    if val is not None:
                        self._write_data_row(ws_stats, row_idx, [label, round(val, 3) if isinstance(val, float) else val])
                        row_idx += 1

            ws_stats.column_dimensions['A'].width = 30
            ws_stats.column_dimensions['B'].width = 15

        return self._save_workbook(wb)

    async def generate_team_report(self, team_id: str, include_players: bool = True, include_stats: bool = True) -> bytes:
        team = await self.fetch_team_data(team_id)
        stats = await self.fetch_team_stats(team_id) if include_stats else {}

        wb = Workbook()
        ws = wb.active
        ws.title = "Team Information"

        ws['A1'] = f"Team Report: {team.get('name', team_id)}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:B1')
        ws['A3'] = "Generated:"
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M')

        self._write_header_row(ws, 5, ['Field', 'Value'])
        info_fields = [
            ('Name', team.get('name') or 'N/A'),
            ('Short Name', team.get('short_name') or team.get('shortName') or 'N/A'),
            ('Country', team.get('country') or 'N/A'),
            ('Stadium', team.get('stadium') or 'N/A'),
            ('Manager', team.get('manager') or 'N/A'),
            ('Founded', team.get('founded') or 'N/A'),
        ]
        for row_idx, (field, value) in enumerate(info_fields, 6):
            self._write_data_row(ws, row_idx, [field, str(value)])

        if include_stats and stats:
            row_start = 6 + len(info_fields) + 2
            self._write_header_row(ws, row_start, ['Statistic', 'Value'])
            stat_labels = {
                'goals': 'Goals Scored',
                'goals_conceded': 'Goals Conceded',
                'wins': 'Wins',
                'draws': 'Draws',
                'losses': 'Losses',
                'matches_played': 'Matches Played',
                'pass_accuracy': 'Pass Accuracy (%)',
                'total_pass': 'Total Passes',
                'clean_sheets': 'Clean Sheets',
            }
            row_idx = row_start + 1
            for key, label in stat_labels.items():
                val = stats.get(key)
                if val is not None:
                    self._write_data_row(ws, row_idx, [label, val])
                    row_idx += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        return self._save_workbook(wb)

    async def generate_match_report(self, match_id: str, include_events: bool = True, include_stats: bool = True) -> bytes:
        match = await self.fetch_match_data(match_id)

        wb = Workbook()
        ws = wb.active
        ws.title = "Match Information"

        home = match.get('home_team_name') or match.get('homeTeamName') or match.get('home_team_id') or 'Home'
        away = match.get('away_team_name') or match.get('awayTeamName') or match.get('away_team_id') or 'Away'

        ws['A1'] = f"Match Report: {home} vs {away}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:B1')
        ws['A3'] = "Generated:"
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M')

        self._write_header_row(ws, 5, ['Field', 'Value'])
        home_score = match.get('home_score', match.get('homeScore', 'N/A'))
        away_score = match.get('away_score', match.get('awayScore', 'N/A'))
        score = f"{home_score} - {away_score}" if home_score != 'N/A' else 'N/A'
        match_fields = [
            ('Match ID', str(match_id)),
            ('Home Team', str(home)),
            ('Away Team', str(away)),
            ('Score', str(score)),
            ('Date', str(match.get('date') or 'N/A')),
            ('Competition', str(match.get('competition') or match.get('competitionID') or 'N/A')),
            ('Venue', str(match.get('venue') or 'N/A')),
            ('Status', str(match.get('status') or 'N/A')),
        ]
        for row_idx, (field, value) in enumerate(match_fields, 6):
            self._write_data_row(ws, row_idx, [field, value])

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30

        return self._save_workbook(wb)
