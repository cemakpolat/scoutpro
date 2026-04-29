"""
Export Handler Service
Handles data fetching and format conversion
"""
import csv
import json
import io
import httpx
import logging
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)


class ExportHandler:
    """Handles data export operations"""

    def __init__(self):
        self.player_service_url = "http://player-service:8000"
        self.team_service_url = "http://team-service:8000"
        self.match_service_url = "http://match-service:8000"
        self.statistics_service_url = "http://statistics-service:8000"
        self.client = httpx.AsyncClient(timeout=30.0)

    async def get_players_data(
        self,
        entity_ids: Optional[List[str]] = None,
        filters: Optional[Dict[str, Any]] = None,
        fields: Optional[List[str]] = None,
        limit: int = 1000
    ) -> List[Dict[str, Any]]:
        try:
            params = {"limit": limit}
            if filters:
                params.update(filters)
                
            # If we want specific IPs, we might need a POST or multiple GETs
            # For simplicity, assuming the player service accepts a list of IDs or we query and filter
            response = await self.client.get(f"{self.player_service_url}/api/v2/players", params=params)
            
            if response.status_code != 200:
                logger.warning(f"Failed to fetch from player-service: {response.text}")
                return []
                
            data = response.json()
            players = data.get("players", data if isinstance(data, list) else [])
            
            if entity_ids:
                players = [p for p in players if p.get('player_id') in entity_ids or p.get('id') in entity_ids]
                
            if fields:
                players = [{k: v for k, v in player.items() if k in fields} for player in players]
                
            return players
        except Exception as e:
            logger.error(f"Error fetching players data from service: {e}")
            return []

    async def get_teams_data(
        self,
        entity_ids: Optional[List[str]] = None,
        filters: Optional[Dict[str, Any]] = None,
        fields: Optional[List[str]] = None,
        limit: int = 1000
    ) -> List[Dict[str, Any]]:
        try:
            params = {"limit": limit}
            if filters:
                params.update(filters)
                
            response = await self.client.get(f"{self.team_service_url}/api/v2/teams", params=params)
            
            if response.status_code != 200:
                logger.warning(f"Failed to fetch from team-service: {response.text}")
                return []
                
            data = response.json()
            teams = data.get("teams", data if isinstance(data, list) else [])

            if entity_ids:
                teams = [t for t in teams if t.get('team_id') in entity_ids or t.get('id') in entity_ids]

            if fields:
                teams = [{k: v for k, v in team.items() if k in fields} for team in teams]

            return teams
        except Exception as e:
            logger.error(f"Error fetching teams data from service: {e}")
            return []

    async def get_matches_data(
        self,
        entity_ids: Optional[List[str]] = None,
        filters: Optional[Dict[str, Any]] = None,
        fields: Optional[List[str]] = None,
        limit: int = 1000
    ) -> List[Dict[str, Any]]:
        try:
            params = {"limit": limit}
            if filters:
                params.update(filters)
                
            response = await self.client.get(f"{self.match_service_url}/api/v2/matches", params=params)
            
            if response.status_code != 200:
                logger.warning(f"Failed to fetch from match-service: {response.text}")
                return []
                
            data = response.json()
            matches = data.get("matches", data if isinstance(data, list) else [])

            if entity_ids:
                matches = [m for m in matches if m.get('match_id') in entity_ids or m.get('id') in entity_ids]

            if fields:
                matches = [{k: v for k, v in match.items() if k in fields} for match in matches]

            return matches
        except Exception as e:
            logger.error(f"Error fetching matches data from service: {e}")
            return []

    async def get_statistics_data(
        self,
        filters: Optional[Dict[str, Any]] = None,
        fields: Optional[List[str]] = None,
        limit: int = 1000
    ) -> List[Dict[str, Any]]:
        try:
            params = {"limit": limit}
            if filters:
                params.update(filters)
                
            response = await self.client.get(f"{self.statistics_service_url}/api/v2/statistics", params=params)
            
            if response.status_code != 200:
                logger.warning(f"Failed to fetch from statistics-service: {response.text}")
                return []
                
            data = response.json()
            statistics = data.get("statistics", data if isinstance(data, list) else [])

            if fields:
                statistics = [{k: v for k, v in stat.items() if k in fields} for stat in statistics]

            return statistics
        except Exception as e:
            logger.error(f"Error fetching statistics data from service: {e}")
            return []

    def to_csv(self, data: List[Dict[str, Any]]) -> bytes:
        if not data:
            return b"No data available"
            
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        
        return output.getvalue().encode('utf-8')

    def to_json(self, data: List[Dict[str, Any]]) -> bytes:
        return json.dumps(data, indent=2).encode('utf-8')

    def to_excel(self, data: List[Dict[str, Any]], sheet_name: str = "Data") -> bytes:
        if not data:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(["No data available"])
            
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
            
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        headers = list(data[0].keys())
        ws.append(headers)
        
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=str(row_data.get(header, "")))
                
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
