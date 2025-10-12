"""
@author: Cem Akpolat
@created by cemakpolat at 2019-10-02

Link to be profited as a code examples: https://xlsxwriter.readthedocs.io/format.html#set_align
"""

import pandas as pd
import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side, PatternFill, colors, Alignment, Color
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter


excel_dir = "../results/playerStatistics.xlsx"


def writeInExcell(eventType, eventDict):

    # Create some Pandas dataframes from some data.
    df1 = pd.DataFrame([eventDict])
    # df2 = pd.DataFrame.from_dict(eventDict)

    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter("../results/playerStatistics.xlsx", engine="xlsxwriter")

    # Write each dataframe to a different worksheet.
    df1.to_excel(writer, sheet_name=eventType)

    workbook = writer.book
    worksheet = writer.sheets[eventType]

    # Add a header format.
    header_format = workbook.add_format(
        {
            "bold": True,
            "text_wrap": True,
            "valign": "top",
            "fg_color": "#D7E4BC",
            "border": 1,
        }
    )
    # center the value inside the cells
    header_format.set_align("center")
    header_format.set_align("vcenter")

    for col_num, value in enumerate(df1.columns.values):
        worksheet.write(0, col_num + 1, value, header_format)

    # adjust the width size of the columns
    worksheet.set_column(0, col_num + 1, 15)

    # # Apply a conditional format to the cell range.
    # worksheet.conditional_format('B2:B8', {'type': '3_color_scale'})

    # # Create a chart object.
    # chart = workbook.add_chart({'type': 'column'})

    # # Configure the series of the chart from the dataframe data.
    # chart.add_series({'values': '=Aerial!$B$2:$B$8'})

    # # Insert the chart into the worksheet.
    # worksheet.insert_chart('D2', chart)

    # Close the Pandas Excel writer and output the Excel file.
    writer.save()
    writer.close()


# https://realpython.com/openpyxl-excel-spreadsheets-python/
def appendResults(eventType, eventDict):
    book = load_workbook(excel_dir)
    with pd.ExcelWriter(excel_dir, engine="openpyxl") as writer:
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        ## Your dataframe to append.
        # Write each dataframe to a different worksheet.
        df1 = pd.DataFrame([eventDict])
        df1.to_excel(writer, sheet_name=eventType)
        sheet = writer.sheets[eventType]

        header = NamedStyle(name="header")
        header.font = Font(bold=True)
        header.fill = PatternFill(
            patternType="solid", fill_type="solid", fgColor=Color("D7E4BC")
        )
        header.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        header.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text="true"
        )

        header_row = sheet[1]

        for cell in header_row:
            cell.style = header
            # sheet.column_dimensions[get_column_letter(cell.column)].width = 15 # this line works too

        # adjust the size of the columns
        dims = {}
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max(
                        (dims.get(cell.column, 0), len(str(cell.value)))
                    )
        for col, value in dims.items():
            sheet.column_dimensions[get_column_letter(col)].width = 15

        writer.save()
        writer.close()


# def readExcel():
#
#     book = xlrd.open_workbook('../results/pandas_multiple.xlsx')
#     print(book.nsheets)
#     print(book.sheet_names())
#     sheet = book.sheet_by_index(1)
#     cell = sheet.cell(0, 1)  # where row=row number and col=column number
#     print (cell.value)  # to print the cell contents
